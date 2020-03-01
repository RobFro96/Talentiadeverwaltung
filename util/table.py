"""
Talentiadeverwaltung für Sportwettkämpfe der Ruderjugend Sachsen
Programm zum Einlesen, Bearbeiten und Abspeichern von Excel-Tabellen
von Robert Fromm (Ruderclub Eilenburg e. V.), Februar 2020
Email: robert_fromm@web.de
"""
import enum
import logging
import numbers
import os
import tkinter.messagebox

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from util.cell import Cell
from util.column_range import ColumnRange

CLOSE_WARNING = "Die Datei %s ist gerade von einem anderen Programm geöffnet. Jetzt schließen?"


class ValueType(enum.Enum):
    """Enum der Datentypen einer Tabelle
    """
    STRING = 0
    NUMBER = 1
    COLUMN_RANGE = 2
    STRING_LIST = 3
    CELL = 4


class Table:
    """Klasse zum Öffnen, Schreiben und Auslesen von Excel-Tabellen mit Hilfe der openpyxl-Bibliothek
    """

    def __init__(self, folder, filename):
        """Konstruktor.

        Args:
            folder (str): Ordner
            filename (str): Dateiname
        """
        self.folder = folder
        self.filename = filename

        self.workbook: Workbook = None
        self.worksheet: Worksheet = None

    def get_path(self):
        """Erstellen des Dateipfades aus Ordner und Dateiname

        Returns:
            str: Dateipfad
        """
        return os.path.join(self.folder, self.filename)

    def file_exists(self):
        """Gibt an, ob die Datei der Tabelle schon existiert.

        Returns:
            bool: True, wenn die Datei existiert
        """
        return os.path.isfile(self.get_path())

    def open(self) -> bool:
        """Öffnen der Tabelle

        Returns:
            bool: True, wenn erfolgreich
        """
        try:
            self.workbook = openpyxl.open(self.get_path())
            self.set_worksheet(0)
        except:
            logging.exception("Fehler beim Öffnen der Tabelle %s.", self.filename)
            return False
        return True

    def get_worksheet(self, index):
        """Gibt das Arbeitsblatt an.
        Wenn index eine Zahl ist, wird die Nummerierung verwendet. Die Indizierung beginnt bei 0.
        Wenn index eine Zeichkette ist, wird nach dem entsprechenden Titel das Arbeitsblattes gesucht.

        Args:
            index (typing.Any): Verweis auf Arbeitsblatt.

        Returns:
            [type]: [description]
        """
        if isinstance(index, int):
            if index >= 0 and index < len(self.workbook.worksheets):
                return self.workbook.worksheets[index]
            else:
                logging.error("Arbeitsblatt %d existiert nicht in der Tabelle %s",
                              index, self.filename)
        elif isinstance(index, str):
            if index in self.workbook.sheetnames:
                return self.workbook[index]
            else:
                logging.error("Arbeitsblatt %s existiert nicht in der Tabelle %s",
                              index, self.filename)
        else:
            logging.error("Unbekannter Datentyp in set_worksheet(%r) der Tabelle %s",
                          index, self.filename)
        return None

    def set_worksheet(self, index):
        """Setzen des angegebenen Arbeitsblattes als aktiv

        Args:
            index (typing.Any): Index entsprechend dem Argument von get_worksheet()
        """
        self.worksheet = self.get_worksheet(index)

    def remove_worksheet(self, index):
        """Entfernen des angegebenen Arbeitsblattes

        Args:
            index (typing.Any): Index entsprechend dem Argument von get_worksheet()
        """
        ws = self.get_worksheet(index)
        if ws:
            self.workbook.remove(ws)
        else:
            logging.error("Arbeitsblatt %s der Tabelle %s konnte nicht entfernt werden.",
                          index, self.filename)

    def copy_worksheet(self, index):
        """Kopieren des angegebenen Arbeitsblattes.
        Die Kopie wird als aktives Arbeitsblatt gesetzt.

        Args:
            index (typing.Any): Index entsprechend dem Argument von get_worksheet()
        """
        ws = self.get_worksheet(index)
        if ws:
            self.worksheet = self.workbook.copy_worksheet(ws)
        else:
            logging.error("Arbeitsblatt %s der Tabelle %s konnte nicht kopiert werden.",
                          index, self.filename)

    def write(self, folder=None, filename=None) -> bool:
        """Schreiben des Tabellenblattes

        Args:
            folder (str, optional): Alternativer Ordner. Defaults to None.
            filename (str, optional): Alternativer Dateiname. Defaults to None.

        Returns:
            bool: True, wenn erfolgreich
        """
        folder = folder or self.folder
        filename = filename or self.filename
        path = os.path.join(folder, filename)

        self.check_lock(folder, filename)

        try:
            self.workbook.save(path)
        except:
            logging.exception("Fehler beim Speichern der Tabelle %s.", filename)
            return False
        return True

    def check_lock(self, folder, filename):
        """Überprüfen, ob die Dateiname in Excel geöffnet ist.
        Es wird nach den Dateien .~lock.___ und ~$___ gesucht.

        Args:
            folder (str): Ordner
            filename (str): Dateiname
        """
        path1 = os.path.join(folder, ".~lock." + filename + "#")
        path2 = os.path.join(folder, "~$" + filename)

        if os.path.exists(path1) or os.path.exists(path2):
            tkinter.messagebox.showwarning("Warnung", CLOSE_WARNING % filename)

    def get_value(self, cell, val_type: ValueType, default=None, log_errors=True):
        """Auslesen eines Tabellenwertes mit angegebenen Typ

        Args:
            cell (typing.Any): Cell beliebigen Types (Cell, String, Tuple)
            val_type (ValueType): Typ
            default (typing.Any, optional): Standardwert. Defaults to None.
            log_errors (bool, optional): True, wenn Fehler angezeigt werden sollen. Defaults to True.

        Returns:
            typing.Any: Wert
        """
        if val_type == ValueType.NUMBER:
            return self.get_number(cell, default, log_errors)
        if val_type == ValueType.COLUMN_RANGE:
            return self.get_col_range(cell, default, log_errors)
        if val_type == ValueType.STRING:
            return self.get_string(cell, default, log_errors)
        if val_type == ValueType.STRING_LIST:
            return self.get_string_list(cell, default, log_errors)
        if val_type == ValueType.CELL:
            return self.get_cell(cell, default, log_errors)
        return default

    @classmethod
    def convert_cell(cls, cell, log_errors=True):
        """Umwandeln des Arguments beliebigen Types in eine Cell

        Args:
            cell (typing.Any): Zelle des Types (str, tuple, Cell)
            log_errors (bool, optional): Anzeige von Fehlermeldungen. Defaults to True.

        Returns:
            Cell: Zelle
        """
        value = None
        if isinstance(cell, str):
            value = Cell.from_string(cell)
        elif isinstance(cell, tuple):
            value = Cell.from_row_col(*cell)
        elif isinstance(cell, Cell):
            value = cell

        if (not value) and log_errors:
            logging.error("Angegebenes Objekt ist keine gültige Zelle: %r.", cell)
        return value

    def get_number(self, cell, default=None, log_errors=True):
        """Auslesen einer Zahl

        Args:
            cell (typing.Any): Zelle beliebigen Types
            default (typing.Any, optional): Standardwert. Defaults to None.
            log_errors (bool, optional): Anzeigen von Fehlermeldungen. Defaults to True.

        Returns:
            int: Zahl
        """
        if default is None:
            default = 0

        cell = self.convert_cell(cell, log_errors)
        if cell is None:
            return default

        value = self.worksheet.cell(*cell.get()).value

        if isinstance(value, numbers.Number):
            return value
        if log_errors:
            logging.error("Zelle %s der Tabelle %s enthält keine Zahl.", cell, self.filename)
        return default

    def get_col_range(self, cell, default=None, log_errors=True):
        """Auslesen eines Spaltenbereiches

        Args:
            cell (typing.Any): Zelle beliebigen Types
            default (typing.Any, optional): Standardwert. Defaults to None.
            log_errors (bool, optional): Anzeigen von Fehlermeldungen. Defaults to True.

        Returns:
            int: Spaltenbereich
        """
        if default is None:
            default = ColumnRange.from_string("A-A")

        cell = self.convert_cell(cell, log_errors)
        if cell is None:
            return default

        value = self.worksheet.cell(*cell.get()).value
        value = ColumnRange.from_string(value)

        if value is not None:
            return value
        if log_errors:
            logging.error("Zelle %s der Tabelle %s enthält keinen Spaltenbereich.",
                          cell, self.filename)
        return default

    def get_string(self, cell, default=None, log_errors=True):
        """Auslesen einer Zeichenkette

        Args:
            cell (typing.Any): Zelle beliebigen Types
            default (typing.Any, optional): Standardwert. Defaults to None.
            log_errors (bool, optional): Anzeigen von Fehlermeldungen. Defaults to True.

        Returns:
            int: Zeichenkette
        """
        if default is None:
            default = ""

        cell = self.convert_cell(cell, log_errors)
        if cell is None:
            return default

        value = self.worksheet.cell(*cell.get()).value

        if value is not None:
            return str(value)
        if log_errors:
            logging.error("Zelle %s der Tabelle %s enthält keine Zeichenkette.",
                          cell, self.filename)
        return default

    def get_string_list(self, cell, default=None, log_errors=True):
        """Auslesen einer Liste von Zeichenkette

        Args:
            cell (typing.Any): Zelle beliebigen Types
            default (typing.Any, optional): Standardwert. Defaults to None.
            log_errors (bool, optional): Anzeigen von Fehlermeldungen. Defaults to True.

        Returns:
            int: Liste von Zeichenketten
        """
        if default is None:
            default = []

        cell = self.convert_cell(cell, log_errors)
        if cell is None:
            return default

        value = self.worksheet.cell(*cell.get()).value

        if value is not None:
            str_list = [element.strip() for element in str(value).split(',')]
            return str_list
        if log_errors:
            logging.error("Zelle %s der Tabelle %s enthält keine Liste.", cell, self.filename)
        return default

    def get_cell(self, cell, default=None, log_errors=True):
        """Auslesen einer Zelle

        Args:
            cell (typing.Any): Zelle beliebigen Types
            default (typing.Any, optional): Standardwert. Defaults to None.
            log_errors (bool, optional): Anzeigen von Fehlermeldungen. Defaults to True.

        Returns:
            int: Zelle
        """
        if default is None:
            default = Cell.from_string("A1")

        cell = self.convert_cell(cell, log_errors)
        if cell is None:
            return default

        value = self.worksheet.cell(*cell.get()).value
        value = Cell.from_string(value)

        if value is not None:
            return value
        if log_errors:
            logging.error("Zelle %s der Tabelle %s enthält keine Zelle.", cell, self.filename)
        return default

    def set_value(self, cell, value, log_errors=True):
        """Setzen eines Wertes in eine Zelele

        Args:
            cell (typing.Any): Zelle beliebigen Types
            value (typing.Any): Wert
            log_errors (bool, optional): Fehleranzeige. Defaults to True.
        """
        cell = self.convert_cell(cell, log_errors)
        if cell is None:
            return

        self.worksheet.cell(*cell.get(), value)

    def set_footer(self, left=None, center=None, right=None, settings=None):
        if not left and settings:
            left = settings["competition_name"]

        if not right and settings:
            right = settings["competition_date"]

        self.worksheet.oddHeader.left.text = left
        self.worksheet.oddHeader.center.text = center
        self.worksheet.oddHeader.right.text = right
